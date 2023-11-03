import openpyxl
from openpyxl import Workbook

dosya = openpyxl.load_workbook("./Enes Şahin - kulucka_ogrenci_listesi.xlsx")
sayfa = dosya["Sayfa1"]
deger = sayfa["A1"].value
dolar_temizleme= str.maketrans({"$": None})
deger_sans_dolar = deger.translate(dolar_temizleme)
print(deger_sans_dolar)
sayfa =dosya["Sayfa1"]
deger = sayfa["B1"].value
dolar_temizleme= str.maketrans({"$": None})
deger_sans_dolar2 = deger.translate(dolar_temizleme)
print(deger_sans_dolar2)
sayfa =dosya["Sayfa1"]
deger = sayfa["C1"].value
dolar_temizleme= str.maketrans({"$": None})
deger_sans_dolar3 = deger.translate(dolar_temizleme)
print(deger_sans_dolar3)
#sayfa.cell(row=2,column=4,value=deger_sans_dolar3)
#sayfa.cell(row=2,column=3,value=deger_sans_dolar2)
#sayfa.cell(row=2,column=1,value=deger_sans_dolar)
sayfa.cell(row=1,column=1, value="İsim")
sayfa.cell(row=1,column=2, value="Soyisim")
sayfa.cell(row=1,column=3, value="Sınıf")
sayfa.cell(row=1,column=4, value="Bölüm")
sayfa.cell(row=2,column=1,value="Yiğit")
sayfa.cell(row=2,column=2,value="Özan")
sayfa.cell(row=2,column=4, value="Elektrik Elektronik Mühendisliği")
sayfa.cell(row=2,column=3, value="2.Sınıf")
sayfa.cell(row=3,column=1, value="Efecan")
sayfa.cell(row=3,column=2, value="Karatut")
sayfa.cell(row=3,column=3, value="2.Sınıf")
sayfa.cell(row=3,column=4, value="Elektrik Elektronik Mühendisliği")
sayfa.cell(row=4,column=1, value="Mustafa")
sayfa.cell(row=4,column=2, value="Çirci")
sayfa.cell(row=4,column=3, value="2.Sınıf")
sayfa.cell(row=4,column=4, value="Bilgisiyar Mühendisliği")
#öncelikle biliyorum "kodu'mun sizin istediğiniz gibi olmadığını fakat gerçekten çok uğraştım elimden geleni yaptım başka bi yerden alarak yapmak yerine öğrendiğim komutlarla yapmak istedim


dosya.save("./Enes Şahin - kulucka_ogrenci_listesi.xlsx")